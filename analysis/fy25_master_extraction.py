"""
FY2025 Master Extraction — Dixstone Internal Metrics
=====================================================
Applies the EXACT April-2026 rev-14 filter set so numbers reconcile to the
daily averages reported in that deck:
  AXIMA $37k/day · BANBA $28k · NUADA $12k · LUG/MIDIR/DRAVUS $21k

Filters applied (per rev-14 methodology):
  ME2N (PO): year=2025, exclude CMT/WIR purchase groups, project cost
             center MUST be empty (non-project / OPEX side only),
             Amount USD populated.

  MB51 (GR): year=2025, Operation Type = WE, project cost center empty,
             year=2025. GR = "OPEX as booked" = SUM of GR movements.

  MB51 (IO): year=2025, movement codes 201/261/281 family, project
             cost center empty.

Outputs to analysis/outputs/:
  fy25_po_by_rig_mg.csv     · long-format: rig, mg, value
  fy25_gr_by_rig_mg.csv
  fy25_io_by_rig_mg.csv
  fy25_totals_by_rig.csv    · per-rig totals for one-pager KPIs

The fleet-wide top-N material groups are computed from these long tables.
"""
import os, csv, time
from collections import defaultdict
import openpyxl

SAP_DIR = ("/Users/ghazi/Library/CloudStorage/OneDrive-SharedLibraries-SigmaEnergy/"
           "Strategy - Documents/Dixstone Rigs - Benchmarking Analysis/"
           "7- SIGMA AI Tools")

OUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
os.makedirs(OUT_DIR, exist_ok=True)

TARGET = {'AXIMA', 'BANBA', 'NUADA', 'LUG', 'MIDIR', 'DRAVUS'}
EXCLUDE_PG = {'CMT', 'WIR'}
OD_2025 = {'AXIMA':310, 'BANBA':365, 'NUADA':365, 'LUG':365, 'MIDIR':365, 'DRAVUS':365}


# ----------------------------------------------------------------------- #
# PO extraction (ME2N)
# ----------------------------------------------------------------------- #
def extract_po():
    print("\n=== PO · ME2N · rev-14 filter ===")
    t0 = time.time()
    wb = openpyxl.load_workbook(os.path.join(SAP_DIR, 'ME2N-rev23.xlsx'),
                                read_only=True, data_only=True)
    ws = wb['RAW']
    by_rig_mg = defaultdict(lambda: defaultdict(float))
    rig_total = defaultdict(float)
    n_rows = 0; n_kept = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        n_rows += 1
        try:
            rig = row[2]; year = row[8]; pg = row[17]
            amt = row[25]; project_cc = row[30]
            mg = row[37]
        except (IndexError, TypeError):
            continue
        if rig not in TARGET: continue
        if year != 2025: continue
        if pg in EXCLUDE_PG: continue
        if project_cc: continue                # KEY: non-project filter
        if not isinstance(amt, (int, float)): continue
        mg_label = (mg or 'NO MATERIAL').strip() if isinstance(mg, str) else 'NO MATERIAL'
        by_rig_mg[rig][mg_label] += amt
        rig_total[rig] += amt
        n_kept += 1
    wb.close()
    print(f"  {n_rows:,} rows · {n_kept:,} kept in {time.time()-t0:.1f}s")
    return by_rig_mg, rig_total


# ----------------------------------------------------------------------- #
# GR / IO extraction (MB51)
# ----------------------------------------------------------------------- #
def extract_mb51(kind):
    """kind = 'GR' or 'IO'"""
    print(f"\n=== {kind} · MB51 · rev-14 filter ===")
    t0 = time.time()
    wb = openpyxl.load_workbook(os.path.join(SAP_DIR, 'MB51-rev71.xlsx'),
                                read_only=True, data_only=True)
    ws = wb['RAW']
    by_rig_mg = defaultdict(lambda: defaultdict(float))
    rig_total = defaultdict(float)
    n_rows = 0; n_kept = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        n_rows += 1
        try:
            rig = row[3]; mg = row[9]; mvt_code = row[17]
            op_type = row[20]; year = row[23]
            amt_abs = row[35] if row[35] is not None else row[34]
            capex_cc = row[51]  # Capex Cost Center (col 52, 0-indexed = 51)
        except (IndexError, TypeError):
            continue
        if rig not in TARGET: continue
        if year != 2025: continue
        if not isinstance(amt_abs, (int, float)): continue
        if capex_cc: continue                  # KEY: non-project (no capex CC)

        if kind == 'GR':
            if op_type != 'WE': continue
        elif kind == 'IO':
            try:
                code = int(mvt_code) if mvt_code is not None else 0
            except (ValueError, TypeError):
                code = 0
            if code not in (201, 261, 281): continue

        mg_label = (mg or 'NO MATERIAL').strip() if isinstance(mg, str) else 'NO MATERIAL'
        v = abs(amt_abs)
        by_rig_mg[rig][mg_label] += v
        rig_total[rig] += v
        n_kept += 1
    wb.close()
    print(f"  {n_rows:,} rows · {n_kept:,} kept in {time.time()-t0:.1f}s")
    return by_rig_mg, rig_total


# ----------------------------------------------------------------------- #
# Persist + report
# ----------------------------------------------------------------------- #
def report(by_rig_mg, rig_total, kind, csv_name):
    """Print per-rig top 5, fleet top 10, and write long CSV."""
    print(f"\n=== {kind} · FY2025 totals ===")
    fleet_total = 0
    fleet_mg = defaultdict(float)
    long_rows = []
    for rig in ['AXIMA','BANBA','NUADA','LUG','MIDIR','DRAVUS']:
        rt = rig_total.get(rig, 0)
        fleet_total += rt
        daily = rt / OD_2025[rig] if OD_2025[rig] else 0
        print(f"  {rig:<8s}  ${rt:>13,.0f}  (${daily/1000:>5.1f}k/day)")
        for mg, v in by_rig_mg.get(rig, {}).items():
            long_rows.append({'rig': rig, 'material_group': mg, 'value_usd': round(v, 2)})
            fleet_mg[mg] += v
    print(f"  {'FLEET':<8s}  ${fleet_total:>13,.0f}")

    # Fleet top 10
    top10 = sorted(fleet_mg.items(), key=lambda x: -x[1])[:10]
    print(f"\n  FLEET TOP-10 material groups · {kind}:")
    for mg, v in top10:
        print(f"    {mg:<18s}  ${v:>13,.0f}  {v/fleet_total*100:>5.1f}%")

    # Write long CSV
    out = os.path.join(OUT_DIR, csv_name)
    with open(out, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['rig','material_group','value_usd'])
        w.writeheader(); w.writerows(long_rows)
    print(f"  Written: {out}")
    return fleet_total, fleet_mg


# ----------------------------------------------------------------------- #
# Main
# ----------------------------------------------------------------------- #
if __name__ == "__main__":
    po_by, po_tot = extract_po()
    po_fleet_total, po_fleet_mg = report(po_by, po_tot, 'PO', 'fy25_po_by_rig_mg.csv')

    gr_by, gr_tot = extract_mb51('GR')
    gr_fleet_total, gr_fleet_mg = report(gr_by, gr_tot, 'GR', 'fy25_gr_by_rig_mg.csv')

    io_by, io_tot = extract_mb51('IO')
    io_fleet_total, io_fleet_mg = report(io_by, io_tot, 'IO', 'fy25_io_by_rig_mg.csv')

    # Per-rig totals for one-pager KPIs
    out = os.path.join(OUT_DIR, 'fy25_totals_by_rig.csv')
    with open(out, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['rig','op_days_2025','po_total_usd','po_per_day_usd',
                    'gr_total_usd','gr_per_day_usd','io_total_usd','io_per_day_usd'])
        for rig in ['AXIMA','BANBA','NUADA','LUG','MIDIR','DRAVUS']:
            od = OD_2025[rig]
            po = po_tot.get(rig, 0); gr = gr_tot.get(rig, 0); io = io_tot.get(rig, 0)
            w.writerow([rig, od, round(po,2), round(po/od,2),
                        round(gr,2), round(gr/od,2), round(io,2), round(io/od,2)])
    print(f"\nWritten: {out}")
    print("\n✅ master extraction complete")
