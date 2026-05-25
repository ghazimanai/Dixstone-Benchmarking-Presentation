"""
Rig One-Pagers — FY 2025 Top-5 by material group
==================================================
For each of 6 Dixstone rigs (AXIMA, BANBA, NUADA, LUG, MIDIR, DRAVUS),
extract the top-5 material groups by value for FY2025 across:

- PO  (purchase orders)   — from ME2N-rev24.xlsx
- GR  (goods receipt)     — from MB51-rev72.xlsx (movement type WE / 101)
- IO  (issue-out)         — from MB51-rev72.xlsx (movement type WA / 201/261)

Filters (mirror the April 2026 internal methodology):
  - Year = 2025
  - Rig in the 6 active rigs
  - Purchasing Group not in (CMT, WIR)  — drilling-only scope

Outputs:
  analysis/outputs/onepagers_po_top5_fy25.csv
  analysis/outputs/onepagers_gr_top5_fy25.csv
  analysis/outputs/onepagers_io_top5_fy25.csv
"""

import os, csv, time
from collections import defaultdict
import openpyxl

SAP_DIR = ("/Users/ghazi/Library/CloudStorage/OneDrive-SharedLibraries-SigmaEnergy/"
           "Strategy - Documents/Dixstone Rigs - Benchmarking Analysis/"
           "7- SIGMA AI Tools")

OUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
os.makedirs(OUT_DIR, exist_ok=True)

TARGET_RIGS = {'AXIMA', 'BANBA', 'NUADA', 'LUG', 'MIDIR', 'DRAVUS'}
EXCLUDE_PG  = {'CMT', 'WIR'}


def extract_po():
    """ME2N · column indexing (0-based, after iter_rows tuple):
       2=Rig, 8=Year, 17=PurchGroup, 25=Amount USD, 37=Material Group"""
    print("\n=== PO (ME2N) — streaming ===")
    t0 = time.time()
    wb = openpyxl.load_workbook(os.path.join(SAP_DIR, 'ME2N-rev24.xlsx'),
                                read_only=True, data_only=True)
    ws = wb['RAW']
    # rig -> mg -> value
    agg = defaultdict(lambda: defaultdict(float))
    rig_total = defaultdict(float)
    n_rows = 0
    n_kept = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        n_rows += 1
        try:
            rig = row[2]; year = row[8]; pg = row[17]; amt = row[25]; mg = row[37]
        except (IndexError, TypeError):
            continue
        if rig not in TARGET_RIGS: continue
        if year != 2025: continue
        if pg in EXCLUDE_PG: continue
        if not isinstance(amt, (int, float)): continue
        mg_label = (mg or 'NO MATERIAL').strip() if isinstance(mg, str) else (mg or 'NO MATERIAL')
        agg[rig][mg_label] += amt
        rig_total[rig] += amt
        n_kept += 1
    wb.close()
    print(f"  scanned {n_rows:,} rows, kept {n_kept:,} (2025, target rigs, ex-CMT/WIR) in {time.time()-t0:.1f}s")
    return agg, rig_total


def extract_mb51(movement_type_filter):
    """MB51 · column indexing (0-based):
       3=Rig, 9=Material Group consolidated, 17=Movement Code2,
       20=Operation Type, 23=Year, 35=Amount USD signed, 36=Amount USD absolute
       For GR keep operation type 'WE' (Goods Receipt).
       For IO keep movement code 261/201/281 family (Issue to cost center).
    """
    print(f"\n=== MB51 — streaming · filter='{movement_type_filter}' ===")
    t0 = time.time()
    wb = openpyxl.load_workbook(os.path.join(SAP_DIR, 'MB51-rev72.xlsx'),
                                read_only=True, data_only=True)
    ws = wb['RAW']
    agg = defaultdict(lambda: defaultdict(float))
    rig_total = defaultdict(float)
    n_rows = 0
    n_kept = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        n_rows += 1
        try:
            rig = row[3]; mg = row[9]; mvt_code = row[17]
            op_type = row[20]; year = row[23]
            amt_abs = row[35] if row[35] is not None else row[34]
        except (IndexError, TypeError):
            continue
        if rig not in TARGET_RIGS: continue
        if year != 2025: continue
        if not isinstance(amt_abs, (int, float)): continue

        # GR filter: Operation Type = WE
        if movement_type_filter == 'GR':
            if op_type != 'WE': continue
        # IO filter: issue-to-cost-center movements (201, 261, 281)
        elif movement_type_filter == 'IO':
            try:
                code_int = int(mvt_code) if mvt_code is not None else 0
            except (ValueError, TypeError):
                code_int = 0
            if code_int not in (201, 261, 281): continue

        mg_label = (mg or 'NO MATERIAL').strip() if isinstance(mg, str) else (mg or 'NO MATERIAL')
        # Use absolute value for IO since they're booked negative
        v = abs(amt_abs) if movement_type_filter == 'IO' else amt_abs
        if v < 0: v = abs(v)
        agg[rig][mg_label] += v
        rig_total[rig] += v
        n_kept += 1
    wb.close()
    print(f"  scanned {n_rows:,} rows, kept {n_kept:,} in {time.time()-t0:.1f}s")
    return agg, rig_total


def top5(agg, rig_total, kind):
    """Print and persist top-5 material groups per rig for the given kind (PO/GR/IO)."""
    out_path = os.path.join(OUT_DIR, f"onepagers_{kind.lower()}_top5_fy25.csv")
    rows = []
    print(f"\n=== TOP 5 · {kind} · FY2025 ===")
    for rig in ['AXIMA', 'BANBA', 'NUADA', 'LUG', 'MIDIR', 'DRAVUS']:
        total = rig_total[rig]
        items = sorted(agg[rig].items(), key=lambda x: -x[1])[:5]
        print(f"\n  {rig}   (total ${total:,.0f}):")
        for mg, v in items:
            pct = v/total*100 if total else 0
            print(f"    {mg:<20s} ${v:>14,.0f}  {pct:>5.1f}%")
            rows.append({"rig": rig, "kind": kind, "material_group": mg,
                         "value_usd": round(v, 2), "rig_total_usd": round(total, 2),
                         "pct_of_rig_total": round(pct, 2)})

    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()) if rows else ['rig','kind','material_group','value_usd','rig_total_usd','pct_of_rig_total'])
        w.writeheader()
        w.writerows(rows)
    print(f"\n  Written: {out_path}")
    return rows


if __name__ == "__main__":
    po_agg, po_total = extract_po()
    top5(po_agg, po_total, 'PO')

    gr_agg, gr_total = extract_mb51('GR')
    top5(gr_agg, gr_total, 'GR')

    io_agg, io_total = extract_mb51('IO')
    top5(io_agg, io_total, 'IO')

    print("\n✅ done")
