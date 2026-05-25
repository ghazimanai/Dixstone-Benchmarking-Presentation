"""
Slide 6 — Peer CAPEX / Revenue · FY 2025
=========================================
Computes the comparable CAPEX / Revenue ratio for each drilling-contractor
peer from the latest extraction in:
  Peer Benchmarking/Drilling Contractors Documents/drilling_contractors_financial_metrics.xlsx

Outputs:
- analysis/outputs/slide6_peer_capex_fy25.csv
- Console print of the peer-median + best/worst by segment

The values feed slide 6 (Segment View) of the executive deck.

Run:
    python3 analysis/slide6_peer_capex_fy25.py
"""

import os
import csv
from openpyxl import load_workbook

SOURCE = ("/Users/ghazi/Library/CloudStorage/OneDrive-SharedLibraries-SigmaEnergy/"
          "Strategy - Documents/Dixstone Rigs - Benchmarking Analysis/"
          "7- SIGMA AI Tools/Peer Benchmarking/Drilling Contractors Documents/"
          "drilling_contractors_financial_metrics.xlsx")

OUT_CSV = os.path.join(os.path.dirname(__file__), "outputs", "slide6_peer_capex_fy25.csv")


# (peer label, sheet name, FY25 revenue row, FY25 capex rows[list], unit-multiplier, segment)
# Unit multiplier converts to USD millions for consistent reporting.
# CAPEX rows lists may include multiple sources (additions + deposits for acquisitions)
# to capture full capital deployment in the year.
PEERS = [
    # Premium-JU heavy peers
    ("Borr Drilling",      "Borr Drilling_raw",        12, [95, 97, 98], 1.0,   "Premium JU"),       # purchases + deposits + additions
    ("Valaris (Jackups)",  "Valaris Jackups+ARO",       9, [19],         1.0,   "Premium/Standard JU"),
    ("ADES",               "ADES_raw",                  8, [],           0.001, "Blended JU+Land"),  # no direct CAPEX line extracted yet
    ("Shelf Drilling",     "Shelf Drilling_raw",       10, [84],         1.0,   "Standard JU"),      # FY25 H1 only — annualise w/ caveat
    ("Velesto",            "Velesto_raw",               8, [],           0.001, "Standard JU"),      # no CAPEX line extracted yet
    # Land/platform peers
    ("Helmerich & Payne",  "H&P_raw",                  10, [97],         0.001, "Land + Platform"),
    ("Nabors Industries",  "Nabors_raw",               10, [103],        0.001, "Land + Platform"),
    ("Precision Drilling", "Precision Drilling_raw",    8, [86],         0.001, "Land"),
]


def fy25_value(ws, row):
    """FY2025 figure is always in column 2 (col B) of these tabs."""
    if row is None:
        return None
    v = ws.cell(row, 2).value
    return abs(v) if isinstance(v, (int, float)) else None


def fy25_capex_total(ws, rows):
    """Sum all CAPEX-related rows for the year (additions, deposits, etc.)."""
    total = 0
    for r in rows:
        v = fy25_value(ws, r)
        if v:
            total += v
    return total if total else None


def main():
    wb = load_workbook(SOURCE, data_only=True)
    rows = []
    for label, sheet, rev_row, capex_rows, mult, segment in PEERS:
        ws = wb[sheet]
        rev = fy25_value(ws, rev_row)
        capex = fy25_capex_total(ws, capex_rows)
        rev_m = rev * mult if rev else None
        capex_m = capex * mult if capex else None
        ratio = (capex_m / rev_m * 100) if (rev_m and capex_m) else None
        rows.append({
            "peer": label, "segment": segment,
            "revenue_fy25_usdm": round(rev_m, 1) if rev_m else None,
            "capex_fy25_usdm":  round(capex_m, 1) if capex_m else None,
            "capex_pct_revenue_fy25": round(ratio, 1) if ratio else None,
        })

    # Console output, segmented
    print(f"{'Peer':<22s} {'Segment':<22s} {'Rev $M':>10s} {'CAPEX $M':>10s} {'CAPEX/Rev':>10s}")
    print("-" * 80)
    for r in rows:
        print(f"{r['peer']:<22s} {r['segment']:<22s} "
              f"{r['revenue_fy25_usdm'] or '-':>10} "
              f"{r['capex_fy25_usdm'] or '-':>10} "
              f"{(str(r['capex_pct_revenue_fy25'])+'%' if r['capex_pct_revenue_fy25'] else '-'):>10}")

    # Persist
    os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
    with open(OUT_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"\nWritten: {OUT_CSV}")

    # Segment-level medians used in slide 6
    print("\n=== Segment medians for slide 6 (used as peer-median benchmark) ===")
    for seg_key, seg_match in [
        ("Premium JU",       ["Borr", "Valaris"]),
        ("Standard JU",      ["Shelf", "Valaris", "Velesto"]),
        ("Land + Platform",  ["Helmerich", "Nabors", "Precision"]),
    ]:
        vals = [r["capex_pct_revenue_fy25"] for r in rows
                if any(s in r["peer"] for s in seg_match)
                and r["capex_pct_revenue_fy25"] is not None]
        if vals:
            vals.sort()
            mid = vals[len(vals)//2] if len(vals) % 2 else (vals[len(vals)//2 - 1] + vals[len(vals)//2]) / 2
            print(f"  {seg_key:<18s}  values={vals}  median={mid}%")


if __name__ == "__main__":
    main()
