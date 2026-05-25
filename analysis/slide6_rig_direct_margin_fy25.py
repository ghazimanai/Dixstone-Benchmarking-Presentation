"""
Slide 6 — Rig Direct Margin normalized · FY 2025
=================================================
Computes the canonical Rig Direct Margin normalized % for Dixstone (per rig)
and the drilling-contractor peer set, sourced from:
  Peer Benchmarking/Drilling Contractors Documents/drilling_contractors_financial_metrics.xlsx

Definition (per normalization_rules.md §2.0):
    Rig Direct Margin normalized =
        Revenue (drilling, excluding reimbursable, bareboat, management, RP)
      − Cost of Revenue normalized
    where Cost of Revenue normalized excludes D&A, SG&A, financial items,
    taxes, and includes OPEX reclassed from CAPEX.

The peer FY2025 values come directly from the per-peer normalization
sections of normalization_rules.md (each peer's section in §3.x states
the final calculated values in a summary table).

Outputs:
- analysis/outputs/slide6_rig_direct_margin_fy25.csv
- Console print used to populate slide 6 cells.
"""

import os, csv
from openpyxl import load_workbook

SOURCE = ("/Users/ghazi/Library/CloudStorage/OneDrive-SharedLibraries-SigmaEnergy/"
          "Strategy - Documents/Dixstone Rigs - Benchmarking Analysis/"
          "7- SIGMA AI Tools/Peer Benchmarking/Drilling Contractors Documents/"
          "drilling_contractors_financial_metrics.xlsx")

OUT_CSV = os.path.join(os.path.dirname(__file__), "outputs", "slide6_rig_direct_margin_fy25.csv")


def main():
    wb = load_workbook(SOURCE, data_only=True)

    # --- Dixstone per-rig: pulled live from Dixstone_P&L tab ---
    ws = wb['Dixstone_P&L']
    rig_cols = {ws.cell(4, c).value: c for c in range(2, 14) if ws.cell(4, c).value not in (None, '%')}
    dix = []
    for rig, col in rig_cols.items():
        rev = ws.cell(10, col).value
        margin = ws.cell(35, col).value   # Rig Direct Margin normalized
        pct = (margin / rev * 100) if (rev and margin is not None) else None
        category = ws.cell(3, col).value
        dix.append({"entity": rig, "segment": category, "revenue_usdk": rev,
                    "rig_direct_margin_usdk": margin,
                    "rig_direct_margin_pct": round(pct, 1) if pct is not None else None,
                    "source": "Dixstone_P&L tab, row 35"})

    # --- Peers: numbers from normalization_rules.md §3.x summary tables ---
    # These are FY2025 values from each peer's per-peer normalization section.
    peer_data = [
        ("Borr Drilling",      "Premium JU (pure)",         906.7,   500.6,  44.8, "normalization_rules.md §3.2"),
        ("ADNOC Drilling",     "Mixed (land + offshore)",  4902.886, 2612.899, 46.7, "normalization_rules.md §3.1B"),
        ("ADES Holding",       "Mixed JU + onshore + barge",  1783.723, 700.878, 60.7, "normalization_rules.md §3.1"),
        ("Valaris (Jackups)",  "JU segment",                  823.4,   486.5,  40.9, "normalization_rules.md §3.7 segment table"),
        ("Valaris (ARO JV)",   "JU (Saudi Aramco JV)",        571.0,   360.7,  36.8, "normalization_rules.md §3.7 segment table"),
        ("Shelf Drilling",     "Standard JU",                 470.2,   259.0,  44.9, "normalization_rules.md §3.6 — H1 2025 only"),
        ("Velesto Energy",     "Standard JU (Malaysia)",      212.612, 108.367,49.0, "normalization_rules.md §3.9"),
        ("Helmerich & Payne",  "Land + Platform (US)",       3678.660,2511.408,31.7, "normalization_rules.md §3.3"),
        ("Nabors Industries",  "Land + Platform (Global)",   3184.693,1914.376,39.9, "normalization_rules.md §3.4"),
        ("Precision Drilling", "Land (Canada + US Intl)",    1316.931, 885.464,32.8, "normalization_rules.md §3.5"),
        ("Vantage Drilling",   "Drillship (divestment year)",  79.180,  85.839, -8.4, "normalization_rules.md §3.8 — anomaly FY25"),
    ]

    print(f"\n{'Entity':<22s} {'Segment':<32s} {'Rev':>10s} {'Cost':>10s} {'Margin %':>10s}")
    print("-" * 95)
    print("=== DIXSTONE (per rig) ===")
    for d in dix:
        rev = d['revenue_usdk'] / 1000  # to $M for readable display
        m = d['rig_direct_margin_pct']
        print(f"  {d['entity']:<20s} {d['segment']:<32s} {rev:>10.1f} {'':>10s} {m:>9.1f}%")

    print("\n=== PEERS (FY2025) ===")
    rows = list(dix)
    for name, segment, rev, cost, margin_pct, source in peer_data:
        rows.append({"entity": name, "segment": segment,
                     "revenue_usdm": rev, "cost_of_rev_norm_usdm": cost,
                     "rig_direct_margin_pct": margin_pct,
                     "source": source})
        print(f"  {name:<20s} {segment:<32s} {rev:>10.1f} {cost:>10.1f} {margin_pct:>9.1f}%")

    # --- Segment medians (used as benchmark on slide 6) ---
    print("\n=== PEER MEDIANS BY SEGMENT — these feed slide 6 ===")
    segments = {
        "Premium JU peers (AXIMA comparators)": ["Borr Drilling", "ADNOC Drilling", "ADES Holding"],
        "Standard JU peers (BANBA/NUADA comparators)": ["Shelf Drilling", "Velesto Energy", "Valaris (Jackups)", "Valaris (ARO JV)"],
        "Land peers (MIDIR/DRAVUS comparators)": ["Helmerich & Payne", "Nabors Industries", "Precision Drilling"],
        "Platform peers (LUG comparators)": ["Helmerich & Payne", "Nabors Industries", "Precision Drilling"],
    }
    for label, names in segments.items():
        vals = sorted([r["rig_direct_margin_pct"] for r in rows if r.get("entity") in names])
        if vals:
            mid = vals[len(vals)//2] if len(vals) % 2 else (vals[len(vals)//2-1] + vals[len(vals)//2])/2
            print(f"  {label}")
            print(f"    values: {vals}   median: {mid:.1f}%   best: {max(vals):.1f}%")

    # --- Persist CSV ---
    os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
    with open(OUT_CSV, "w", newline="") as f:
        all_keys = set()
        for r in rows: all_keys.update(r.keys())
        keys = ["entity","segment","revenue_usdk","revenue_usdm","rig_direct_margin_usdk",
                "cost_of_rev_norm_usdm","rig_direct_margin_pct","source"]
        w = csv.DictWriter(f, fieldnames=keys, extrasaction='ignore')
        w.writeheader()
        w.writerows(rows)
    print(f"\nWritten: {OUT_CSV}")


if __name__ == "__main__":
    main()
