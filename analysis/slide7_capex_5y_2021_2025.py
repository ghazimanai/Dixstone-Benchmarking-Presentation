"""
Slide 7 · CAPEX / Theoretical Revenue · 5-year rolling (2021-2025)
==================================================================
User's idea: single-year CAPEX/Rev is too lumpy because of SPS cycles
and one-off equipment replacements. A 5-year sum-of-CAPEX divided by
sum-of-Theoretical-Revenue is the more representative sustaining
capital-intensity metric per rig.

Calculation method:
  CAPEX_5y / Rev_5y = sum(CAPEX_yr) / sum(TheoRev_yr)  for yr in 2021..2025

For each rig, reads from Dixstone_P&L tab:
  Row 10: Theoretical Annual Revenue (USD '000s)
  Row 38: CAPEX (USD '000s)
across columns corresponding to each year block.

Year block column layout in Dixstone_P&L:
  2025: cols 2-15     2024: cols 17-30    2023: cols 32-45
  2022: cols 47-62    2021: cols 64-90

Each year-block has 2 columns per rig (value + %).

Peer comparators (5-year medians from non-financial Excel summary,
verified via user-provided table):
  Premium JU peers (Borr, ADNOC, ADES):  median 25.7%, top ADES 25.7%
  Standard JU peers (Shelf, Velesto, Valaris-JU+ARO): median 19.8%, top Valaris 20.6%
  Land peers (H&P, Nabors, Precision):   median 12.2%, top Nabors 18.0%
  Platform peers (same as land):         median 12.2%, top Nabors 18.0%

Output: analysis/outputs/slide7_capex_5y_2021_2025.csv
"""

import os, csv
from openpyxl import load_workbook

SRC = ("/Users/ghazi/Library/CloudStorage/OneDrive-SharedLibraries-SigmaEnergy/"
       "Strategy - Documents/Dixstone Rigs - Benchmarking Analysis/"
       "7- SIGMA AI Tools/drilling_contractors_financial_metrics.xlsx")

OUT_CSV = os.path.join(os.path.dirname(__file__), "outputs",
                      "slide7_capex_5y_2021_2025.csv")

# Inclusive year-block column ranges. NOTE: 2021 has DAGDA which extends
# the block to col 87, but the search is safe up to 90.
YEAR_RANGES = {
    2025: (2, 15),
    2024: (17, 30),
    2023: (32, 45),
    2022: (47, 62),
    2021: (64, 90),
}

TARGET_RIGS = ['AXIMA', 'BANBA', 'NUADA', 'LUG', 'MIDIR', 'DRAVUS']


def find_rig_cols(ws, year_start, year_end):
    rig_cols = {}
    for c in range(year_start, year_end + 1):
        v = ws.cell(4, c).value
        if v in TARGET_RIGS:
            rig_cols[v] = c
    return rig_cols


def main():
    wb = load_workbook(SRC, data_only=True, read_only=True)
    ws = wb['Dixstone_P&L']

    rig_totals = {r: {'rev': 0, 'capex': 0, 'per_year': {}} for r in TARGET_RIGS}

    for year in sorted(YEAR_RANGES):
        rs, re_ = YEAR_RANGES[year]
        rig_cols = find_rig_cols(ws, rs, re_)
        for rig, col in rig_cols.items():
            rev = ws.cell(10, col).value or 0
            cap = ws.cell(38, col).value or 0
            if not isinstance(rev, (int, float)): rev = 0
            if not isinstance(cap, (int, float)): cap = 0
            rig_totals[rig]['rev'] += rev
            rig_totals[rig]['capex'] += cap
            rig_totals[rig]['per_year'][year] = (rev, cap, (cap/rev*100 if rev else 0))

    # Print + persist
    print(f"\n=== 5-YEAR CAPEX/Rev · 2021-2025 · Dixstone fleet ===\n")
    print(f"{'Rig':<8s} {'5Y Rev ($M)':>14s} {'5Y CAPEX ($M)':>16s} {'5Y CAPEX/Rev':>14s}")
    print("-" * 65)
    rows = []
    for rig in TARGET_RIGS:
        t = rig_totals[rig]
        pct = t['capex'] / t['rev'] * 100 if t['rev'] else 0
        print(f"{rig:<8s} {t['rev']/1000:>14,.1f} {t['capex']/1000:>16,.2f} {pct:>13.1f}%")
        rows.append({
            "rig": rig,
            "rev_5y_usdk": round(t['rev'], 2),
            "capex_5y_usdk": round(t['capex'], 2),
            "capex_pct_rev_5y": round(pct, 2),
            "per_year_capex_pct_rev": "; ".join(
                f"{y}: rev={v[0]:.0f}, capex={v[1]:.0f}, pct={v[2]:.1f}%"
                for y, v in sorted(t['per_year'].items()))
        })

    print(f"\n=== Peer 5Y medians (used as comparators on slide 7) ===")
    print(f"  Premium JU peers (Borr 17.8 · ADNOC 25.7 · ADES 25.7) → median 25.7%, top ADES")
    print(f"  Standard JU peers (Shelf 8.9 · Velesto 19.8 · Valaris 20.6) → median 19.8%, top Valaris")
    print(f"  Land peers (H&P 12.2 · Nabors 18.0 · Precision 11.4) → median 12.2%, top Nabors")

    os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
    with open(OUT_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print(f"\nWritten: {OUT_CSV}")


if __name__ == "__main__":
    main()
